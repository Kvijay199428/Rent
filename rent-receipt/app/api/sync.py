from fastapi import APIRouter, HTTPException, UploadFile, File, Form, BackgroundTasks
from app.core.routes import Names
from fastapi.responses import StreamingResponse, FileResponse
from app.core.dependencies import config
from app.models.tenant import Tenant
import os
import io
import json
import datetime
import zipfile
import csv
import uvicorn
import socket

from app.services.tenant_service import load_tenants, add_tenant, update_tenant
from app.services.billing_service import get_all_receipts
from app.services.backup_service import create_full_backup
from app.core.paths import BACKUPS_DIR

from app.authentication.common.utils import validate_tenant_pin, hash_pin
from app.authentication.common.pin_vault import encrypt_admin_view_pin
from app.authentication.tenant.sessions import revoke_all_tenant_sessions
from app.core.db import get_conn

import openpyxl
from openpyxl.styles import Font, PatternFill
router = APIRouter()


# ==========================================
# EXCEL IMPORT & EXPORT ENGINE (RELATIONAL)
# ==========================================

PROFILE_HEADERS = [
    "Tenant_ID", "Tenant_Name", "Phone", "Email", "Company", "Address", "Room",
    "Meter_ID", "PIN", "Rent", "Water", "Electricity_Rate", "Additional_Person_Rate",
    "Tank_Water", "Status"
]

RECEIPT_HEADERS = [
    "Bill_No", "Tenant_ID", "Month", "Date", "Previous", "Current", "Units", "Rent",
    "Water", "Electricity", "Additional", "Tank_Water", "Maintenance", "Arrears",
    "Amount_Received", "Total", "Payment_Status", "Receipt_Status"
]

def _build_excel_workbook(tenants_list, receipts_list):
    """Shared helper: builds and returns an openpyxl workbook in memory."""
    wb = openpyxl.Workbook()
    ws_profile = wb.active
    ws_profile.title = "Tenant_Profile"
    ws_profile.append(PROFILE_HEADERS)

    ws_receipts = wb.create_sheet("Rent_Receipts")
    ws_receipts.append(RECEIPT_HEADERS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for ws in [ws_profile, ws_receipts]:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    tenant_id_map = {}
    for t in tenants_list:
        t_id_str = f"T{str(t.id).zfill(3)}"
        tenant_id_map[t.name] = t_id_str
        ws_profile.append([
            t_id_str, t.name, str(t.phone), getattr(t, 'email', ''), getattr(t, 'company', ''),
            getattr(t, 'address', ''), getattr(t, 'room_number', ''), getattr(t, 'meter_id', ''),
            getattr(t, 'tenant_pin', ''), float(t.rent), float(t.water), float(t.electricity_rate),
            float(t.additional_person_charge), float(getattr(t, 'default_tank_water_charge', 0.0)), t.status
        ])

    for r in receipts_list:
        t_name = r.get("Tenant", "")
        t_id_str = tenant_id_map.get(t_name, "UNKNOWN")
        ws_receipts.append([
            r.get("Bill", ""), t_id_str, r.get("Month", ""), r.get("Date", ""),
            float(r.get("Previous", 0)), float(r.get("Current", 0)), float(r.get("Units", 0)),
            float(r.get("Rent", 0)), float(r.get("Water", 0)), float(r.get("Electricity", 0)),
            float(r.get("Additional", 0)), float(r.get("Tank_Water", 0)),
            float(r.get("Maintenance_Charge", 0)), float(r.get("Previous_Arrears", 0)),
            float(r.get("Amount_Received", 0)), float(r.get("Total", 0)),
            r.get("Payment_Status", "PENDING"), r.get("Status", "ACTIVE")
        ])

    return wb


@router.get("/api/export/csv", name=Names.EXPORT_RECEIPTS_CSV)
async def export_receipts_csv(tenants_list: str = "all"):
    tenants = load_tenants()
    receipts = get_all_receipts()

    if tenants_list != "all":
        selected_ids = [int(x) for x in tenants_list.split(",") if x.isdigit()]
        selected_names = [t.name for t in tenants if t.id in selected_ids]
        receipts = [r for r in receipts if r.get("Tenant") in selected_names]

    stream = io.StringIO()
    if receipts:
        writer = csv.DictWriter(stream, fieldnames=receipts[0].keys())
        writer.writeheader()
        writer.writerows(receipts)
    else:
        stream.write("No data found for selected tenants.")

    date_str = datetime.datetime.now().strftime('%Y%m%d')
    filename = f"receipts_export_{date_str}.csv"
    
    response = StreamingResponse(iter([stream.getvalue()]), media_type="text/csv")
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

@router.get("/api/export/zip", name=Names.EXPORT_FULL_ZIP)
async def export_full_zip(tenants_list: str = "all"):
    tenants = load_tenants()
    receipts = get_all_receipts()

    if tenants_list != "all":
        selected_ids = [int(x) for x in tenants_list.split(",") if x.isdigit()]
        selected_names = [t.name for t in tenants if t.id in selected_ids]
        receipts = [r for r in receipts if r.get("Tenant") in selected_names]

    date_str = datetime.datetime.now().strftime('%Y%m%d')
    zip_filename = f"tenant_data_{date_str}.zip"
    zip_path = os.path.join(BACKUPS_DIR, zip_filename)
    os.makedirs(BACKUPS_DIR, exist_ok=True)
    
    from app.services.pdf_service import generate_professional_pdf
    from app.core.config_service import config
    landlord_conf = config.get("landlord", {})

    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        stream = io.StringIO()
        if receipts:
            writer = csv.DictWriter(stream, fieldnames=receipts[0].keys())
            writer.writeheader()
            writer.writerows(receipts)
        zipf.writestr("receipts_data.csv", stream.getvalue())

        for r in receipts:
            tenant_name = r.get("Tenant", "Unknown").replace(" ", "_")
            try:
                formatted_date = datetime.strptime(r.get("Date", ""), "%d %B %Y").strftime("%Y%m%d")
            except Exception:
                formatted_date = r.get("Date", "").replace(" ", "")
                
            custom_filename = f"{tenant_name}_{formatted_date}_{r['Bill']}.pdf"
            status = r.get("Status", "ACTIVE")
            folder = "archive" if status == "ARCHIVED" else "active"
            
            try:
                pdf_stream = generate_professional_pdf(r, landlord_conf)
                zipf.writestr(f"PDFs/{folder}/{custom_filename}", pdf_stream.getvalue())
            except Exception as e:
                print(f"Failed to generate PDF for {r['Bill']}: {e}")

    response = FileResponse(zip_path, media_type="application/zip", filename=zip_filename)
    response.headers["Content-Disposition"] = f'attachment; filename="{zip_filename}"'
    return response

@router.get("/api/sync/template", name=Names.DOWNLOAD_EXCEL_TEMPLATE)
async def download_excel_template():
    wb = openpyxl.Workbook()
    ws_profile = wb.active
    ws_profile.title = "Tenant_Profile"
    ws_profile.append(PROFILE_HEADERS)
    ws_receipts = wb.create_sheet("Rent_Receipts")
    ws_receipts.append(RECEIPT_HEADERS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for ws in [ws_profile, ws_receipts]:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    # Sample data rows
    ws_profile.append(["T001", "John Doe", "9876543210", "john@gmail.com", "ABC Pvt Ltd", "Delhi", "A101", "MTR001", "", 15000, 500, 8.5, 1000, 300, "Active"])
    ws_profile.append(["T002", "Alice Smith", "9988776655", "alice@gmail.com", "XYZ Ltd", "Noida", "B202", "MTR002", "", 18000, 600, 9.0, 1200, 400, "Active"])
    ws_receipts.append(["T1-001", "T001", "January 2026", "01 Jan 2026", 120, 150, 30, 15000, 500, 255, 1000, 300, 0, 0, 17055, 17055, "PAID", "ACTIVE"])
    ws_receipts.append(["T2-001", "T002", "January 2026", "01 Jan 2026", 80, 110, 30, 18000, 600, 270, 0, 400, 0, 0, 19270, 19270, "PAID", "ACTIVE"])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = StreamingResponse(iter([stream.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = 'attachment; filename="Rent_Data_Template.xlsx"'
    return response

@router.get("/api/sync/export/{format}", name=Names.EXPORT_EXCEL_DATA)
async def export_excel_data(format: str):
    tenants = load_tenants()
    receipts = get_all_receipts()

    # Build workbook entirely in RAM
    wb = _build_excel_workbook(tenants, receipts)
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    date_str = datetime.datetime.now().strftime('%Y%m%d')

    if format == "xlsx":
        filename = f"Rent_Data_Export_{date_str}.xlsx"
        response = StreamingResponse(
            iter([excel_stream.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    elif format == "zip":
        zip_stream = io.BytesIO()
        with zipfile.ZipFile(zip_stream, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.writestr(f"Rent_Data_Export_{date_str}.xlsx", excel_stream.getvalue())
        zip_stream.seek(0)
        zip_filename = f"Rent_Data_Archive_{date_str}.zip"
        response = StreamingResponse(
            iter([zip_stream.getvalue()]),
            media_type="application/zip"
        )
        response.headers["Content-Disposition"] = f'attachment; filename="{zip_filename}"'
        return response

    raise HTTPException(status_code=400, detail="Unsupported format. Use 'xlsx' or 'zip'.")

def parse_excel_bytes(file_bytes, filename):
    wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    if "Tenant_Profile" not in wb.sheetnames or "Rent_Receipts" not in wb.sheetnames:
        raise ValueError(f"File '{filename}' is missing required sheets 'Tenant_Profile' and/or 'Rent_Receipts'.")

    ws_profile = wb["Tenant_Profile"]
    ws_receipts = wb["Rent_Receipts"]

    p_headers = [str(cell.value).strip() if cell.value else f"Col{i}" for i, cell in enumerate(ws_profile[1])]
    r_headers = [str(cell.value).strip() if cell.value else f"Col{i}" for i, cell in enumerate(ws_receipts[1])]

    tenants_dict = {}
    for row in ws_profile.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        row_dict = {p_headers[i]: (str(val).strip() if val is not None else "") for i, val in enumerate(row)}
        t_id = row_dict.get("Tenant_ID", "")
        if t_id:
            tenants_dict[t_id] = {"profile": row_dict, "receipts": []}

    for row in ws_receipts.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        row_dict = {r_headers[i]: (str(val).strip() if val is not None else "") for i, val in enumerate(row)}
        t_id = row_dict.get("Tenant_ID", "")
        if t_id in tenants_dict:
            tenants_dict[t_id]["receipts"].append(row_dict)

    return tenants_dict

@router.post("/api/sync/import/preview", name=Names.IMPORT_PREVIEW_DATA)
async def import_preview_data(file: UploadFile = File(...)):
    preview_data = {}
    content = await file.read()
    try:
        if file.filename.endswith('.zip'):
            with zipfile.ZipFile(io.BytesIO(content)) as z:
                for zip_info in z.infolist():
                    if zip_info.filename.endswith('.xlsx'):
                        with z.open(zip_info) as f:
                            preview_data[zip_info.filename] = parse_excel_bytes(f.read(), zip_info.filename)
        elif file.filename.endswith('.xlsx'):
            preview_data[file.filename] = parse_excel_bytes(content, file.filename)
        else:
            raise HTTPException(status_code=400, detail="Only .xlsx or .zip files are supported.")
        return {"status": "success", "files": preview_data}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/api/sync/import/execute", name=Names.IMPORT_EXECUTE_DATA)
async def import_execute_data(
    file: UploadFile = File(...),
    selectedtargets: str = Form(...),
    background_tasks: BackgroundTasks = BackgroundTasks()
):
    selected_list = json.loads(selectedtargets)
    if not selected_list:
        raise HTTPException(status_code=400, detail="No tenants selected for import.")

    content = await file.read()
    sys_tenants = load_tenants()
    sys_receipts = get_all_receipts()

    background_tasks.add_task(create_full_backup, tag="pre_import_excel")

    def execute_import_for_file(file_bytes, filename):
        parsed_data = parse_excel_bytes(file_bytes, filename)
        for t_id, t_data in parsed_data.items():
            target_key = f"{filename}::{t_id}"
            if target_key not in selected_list:
                continue

            p = t_data["profile"]
            t_name = p.get("Tenant_Name", "").strip()
            if not t_name:
                continue

            t = next((x for x in sys_tenants if x.name.lower() == t_name.lower()), None)
            is_new = False
            if not t:
                t = Tenant(name=t_name, phone=p.get("Phone", ""), rent=0.0, water=0.0, electricity_rate=0.0)
                is_new = True

            t.phone = p.get("Phone", t.phone)
            t.email = p.get("Email", getattr(t, 'email', ''))
            t.company = p.get("Company", getattr(t, 'company', ''))
            t.address = p.get("Address", getattr(t, 'address', ''))
            t.room_number = p.get("Room", getattr(t, 'room_number', ''))
            t.meter_id = p.get("Meter_ID", getattr(t, "meter_id", ""))

            # --------------------------------------------------
            # Secure Tenant PIN Import
            # --------------------------------------------------

            plain_pin = str(p.get("PIN") or "").strip()

            pin_changed = False
            hashed_pin = None
            encrypted_pin = None

            if plain_pin:
                validate_tenant_pin(plain_pin)

                hashed_pin = hash_pin(plain_pin)
                encrypted_pin = encrypt_admin_view_pin(plain_pin)

                t.tenant_pin = hashed_pin
                pin_changed = True

            t.rent = float(p.get("Rent", t.rent) or 0.0)
            t.water = float(p.get("Water", t.water) or 0.0)
            t.electricity_rate = float(p.get("Electricity_Rate", t.electricity_rate) or 0.0)
            t.additional_person_charge = float(p.get("Additional_Person_Rate", getattr(t, 'additional_person_charge', 0.0)) or 0.0)
            t.default_tank_water_charge = float(p.get("Tank_Water", getattr(t, 'default_tank_water_charge', 0.0)) or 0.0)
            t.status = p.get("Status", getattr(t, 'status', 'Active'))

            if is_new:
                tenant_id = add_tenant(t)
                t.id = tenant_id
                sys_tenants.append(t)

                if pin_changed:
                    now = datetime.datetime.utcnow().isoformat()
                    with get_conn() as conn:
                        conn.execute(
                            """
                            INSERT INTO tenant_pin_history
                            (tenant_id, pin_hash, changed_at)
                            VALUES (?, ?, ?)
                            """,
                            (tenant_id, hashed_pin, now)
                        )
                        conn.execute(
                            """
                            INSERT OR REPLACE INTO tenant_pin_admin_store
                            (tenant_id, encrypted_pin, updated_at)
                            VALUES (?, ?, ?)
                            """,
                            (tenant_id, encrypted_pin, now)
                        )
                        conn.commit()
            else:
                update_tenant(t)

                if pin_changed:
                    now = datetime.datetime.utcnow().isoformat()
                    with get_conn() as conn:
                        conn.execute(
                            """
                            INSERT INTO tenant_pin_history
                            (tenant_id, pin_hash, changed_at)
                            VALUES (?, ?, ?)
                            """,
                            (t.id, hashed_pin, now)
                        )
                        conn.execute(
                            """
                            INSERT OR REPLACE INTO tenant_pin_admin_store
                            (tenant_id, encrypted_pin, updated_at)
                            VALUES (?, ?, ?)
                            """,
                            (t.id, encrypted_pin, now)
                        )
                        conn.commit()
                    # Force tenant to login again using new PIN
                    revoke_all_tenant_sessions(t.id)

            for r in t_data["receipts"]:
                bill_no = r.get("Bill_No", "").strip()
                if not bill_no:
                    continue
                sys_r = next((x for x in sys_receipts if x.get("Bill") == bill_no), None)
                data = {
                    "Bill": bill_no, "Date": r.get("Date", ""), "Month": r.get("Month", ""),
                    "Tenant": t_name, "Previous": r.get("Previous", 0), "Current": r.get("Current", 0),
                    "Units": r.get("Units", 0), "Rent": r.get("Rent", 0), "Additional": r.get("Additional", 0),
                    "Water": r.get("Water", 0), "Tank_Water": r.get("Tank_Water", 0),
                    "Electricity": r.get("Electricity", 0), "Maintenance_Charge": r.get("Maintenance", 0),
                    "Maintenance_Desc": r.get("Maintenance_Desc", ""),
                    "Previous_Arrears": r.get("Arrears", 0), "Amount_Received": r.get("Amount_Received", 0),
                    "Total": r.get("Total", 0), "Payment_Status": r.get("Payment_Status", "PENDING"),
                    "Status": r.get("Receipt_Status", "ACTIVE"), "Receipt_Version": 8, "Generated_By": "Import"
                }
                if sys_r:
                    sys_r.update(data)
                else:
                    sys_receipts.append(data)

    # try:
    #     if file.filename.endswith('.zip'):
    #         with zipfile.ZipFile(io.BytesIO(content)) as z:
    #             for zip_info in z.infolist():
    #                 if zip_info.filename.endswith('.xlsx'):
    #                     with z.open(zip_info) as f:
    #                         execute_import_for_file(f.read(), zip_info.filename)
    #     elif file.filename.endswith('.xlsx'):
    #         execute_import_for_file(content, file.filename)

    #     from app.services.billing_service import save_all_receipts
    #     save_all_receipts(sys_receipts)
    #     return {"status": "success"}
    # except Exception as e:
    #     raise HTTPException(status_code=400, detail=str(e))
    try:
        if file.filename.endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(content)) as z:
                for zip_info in z.infolist():
                    if zip_info.filename.endswith(".xlsx"):
                        with z.open(zip_info) as f:
                            execute_import_for_file(
                                f.read(),
                                zip_info.filename
                            )

        elif file.filename.endswith(".xlsx"):
            execute_import_for_file(content, file.filename)

        from app.services.billing_service import save_all_receipts
        save_all_receipts(sys_receipts)

        return {
            "status": "success",
            "message": "Import completed successfully."
        }

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    finally:
        # Close the uploaded file
        try:
            await file.close()
        except Exception:
            pass

        # Remove Starlette/FastAPI temporary upload file (if one exists)
        try:
            temp_path = getattr(file.file, "name", None)

            if (
                isinstance(temp_path, str)
                and os.path.isfile(temp_path)
            ):
                os.remove(temp_path)

        except Exception:
            # Ignore cleanup errors
            pass


if __name__ == "__main__":
    sys_conf = config.get("system", {})
    server_host = sys_conf["server"]["host"]
    server_port = sys_conf["server"]["port"]
    is_debug = sys_conf["server"]["debug"]
    
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('10.255.255.255', 1))
        local_ip = s.getsockname()[0]
    except Exception:
        local_ip = '127.0.0.1'
    finally:
        s.close()
        
    print(f"\n{'='*50}")
    print(f" {sys_conf['app']['title']} is starting...")
    print(f"{'='*50}")
    print(f" [Local]:   http://127.0.0.1:{server_port}")
    print(f" [Network]: http://{local_ip}:{server_port}")
    print(f" [Note]:    Do NOT click the {server_host} link below")
    print(f"{'='*50}\n")

    uvicorn.run(
        "app:app",
        host=server_host,
        port=server_port,
        reload=is_debug,
        proxy_headers=True,
        forwarded_allow_ips="*",
        access_log=True
    )

