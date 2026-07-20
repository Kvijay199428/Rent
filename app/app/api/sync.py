# File: app/app/api/sync.py
# POLICY: tenantId is the only identity key for tenant-related data.
# tenantName is display-only and must never be used for joins, ownership, lookup, or mutation.
from typing import Optional
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, BackgroundTasks

from app.core.routes_manifest import Names, Routes

from fastapi.responses import StreamingResponse, FileResponse
from app.core.dependencies import config
from app.models.tenant import Tenant
import os
import io
import json
import datetime
import zipfile
import csv
from typing import List
import uvicorn
import socket
import uuid

from app.services.tenant_service import load_tenants, add_tenant, update_tenant
from app.services.billing_service import get_all_receipts
from app.services.backup_service import create_full_backup
from app.core.paths import BACKUPS_DIR

from app.authentication.common.utils import validate_tenantPin, hash_pin
from app.authentication.common.pin_vault import encrypt_admin_view_pin
from app.authentication.tenant.sessions import revoke_all_tenant_sessions
from app.core.db import get_conn

import openpyxl
from openpyxl.styles import Font, PatternFill
router = APIRouter()


# ==========================================
# EXCEL IMPORT & EXPORT ENGINE (RELATIONAL)
# ==========================================

PROFILE_HEADERS = [
    "tenantId", "tenantName", "Phone", "Email", "Company", "Address", "Room",
    "meterId", "PIN", "Rent", "Water", "electricityRate", "additionalPersonRate",
    "tankWater", "Status"
]

RECEIPT_HEADERS = [
    "BillNo", "tenantId", "Month", "Date", "Previous", "Current", "Units", "Rent",
    "Water", "Electricity", "Additional", "tankWater", "Maintenance", "Arrears",
    "amountReceived", "Total", "paymentStatus", "receiptStatus"
]

def _build_excel_workbook(tenants_list, receipts_list):
    """Shared helper: builds and returns an openpyxl workbook in memory."""
    from app.authentication.common.pin_vault import decrypt_admin_view_pin
    from app.core.db import get_conn

    # Pre-fetch all decrypted PINs from admin store
    decrypted_pins = {}
    try:
        with get_conn() as conn:
            rows = conn.execute("SELECT tenantId, encrypted_pin FROM tenantPin_admin_store").fetchall()
            for row in rows:
                try:
                    decrypted_pins[row["tenantId"]] = decrypt_admin_view_pin(row["encrypted_pin"])
                except Exception:
                    decrypted_pins[row["tenantId"]] = ""
    except Exception:
        pass

    wb = openpyxl.Workbook()
    ws_profile = wb.active
    ws_profile.title = "Tenant_Profile"
    ws_profile.append(PROFILE_HEADERS)

    ws_receipts = wb.create_sheet("Rent_Receipts")
    ws_receipts.append(RECEIPT_HEADERS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for ws in [ws_profile, ws_receipts]:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    tenantId_map = {}
    for t in tenants_list:
        t_id_str = f"T{str(t.id).zfill(3)}"
        tenantId_map[t.name] = t_id_str
        # Use decrypted PIN for export, fallback to empty string if not available
        plain_pin = decrypted_pins.get(t.id, "")
        ws_profile.append([
            t_id_str, t.name, str(t.phone), getattr(t, 'email', ''), getattr(t, 'company', ''),
            getattr(t, 'address', ''), getattr(t, 'roomNumber', ''), getattr(t, 'meterId', ''),
            plain_pin, float(t.rent), float(t.water), float(t.electricityRate),
            float(t.additionalPersonCharge), float(getattr(t, 'defaulttankWaterCharge', 0.0)), t.status
        ])

    for r in receipts_list:
        receipt_tid = int(r.get("TenantId", 0) or 0)
        # Use stored TenantId as primary; fall back to name map only for legacy rows
        t_id_str = f"T{str(receipt_tid).zfill(3)}" if receipt_tid else tenantId_map.get(r.get("Tenant", ""), "UNKNOWN")
        ws_receipts.append([
            r.get("Bill", ""), t_id_str, r.get("Month", ""), r.get("Date", ""),
            float(r.get("Previous", 0) or 0), float(r.get("Current", 0) or 0), float(r.get("Units", 0) or 0),
            float(r.get("Rent", 0) or 0), float(r.get("Water", 0) or 0), float(r.get("Electricity", 0) or 0),
            float(r.get("Additional", 0) or 0), float(r.get("tankWater", 0) or 0),
            float(r.get("MaintenanceCharge", 0) or 0), float(r.get("previousArrears", 0) or 0),
            float(r.get("amountReceived", 0) or 0), float(r.get("Total", 0) or 0),
            r.get("paymentStatus", "PENDING"), r.get("Status", "ACTIVE")
        ])

    return wb


@router.get(Routes.ADMINAPISYNCEXPORTCSV, name=Names.EXPORTRECEIPTSCSV)
async def export_receipts_csv(tenants_list: str = "all"):
    tenants = load_tenants()
    receipts = get_all_receipts()

    if tenants_list != "all":
        selected_ids = {int(x) for x in tenants_list.split(",") if x.isdigit()}
        receipts = [r for r in receipts if int(r.get("TenantId", 0) or 0) in selected_ids]

    stream = io.StringIO()
    if receipts:
        writer = csv.DictWriter(stream, fieldnames=receipts[0].keys())
        writer.writeheader()
        writer.writerows(receipts)
    else:
        stream.write("No data found for selected tenants.")

    date_str = datetime.datetime.now().strftime('%Y%m%d')
    filename = f"receipts_export_{date_str}.csv"

    response = StreamingResponse(iter([stream.getvalue()]), media_type="text/csv")
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

@router.get(Routes.ADMINAPISYNCEXPORTZIP, name=Names.EXPORTFULLZIP)
async def export_full_zip(tenants_list: str = "all"):
    tenants = load_tenants()
    receipts = get_all_receipts()

    if tenants_list != "all":
        selected_ids = {int(x) for x in tenants_list.split(",") if x.isdigit()}
        receipts = [r for r in receipts if int(r.get("TenantId", 0) or 0) in selected_ids]

    date_str = datetime.datetime.now().strftime('%Y%m%d')
    zip_filename = f"tenant_data_{date_str}.zip"
    zip_path = os.path.join(BACKUPS_DIR, zip_filename)
    os.makedirs(BACKUPS_DIR, exist_ok=True)

    from app.services.pdf_service import generate_professional_pdf
    from app.core.config_service import config
    landlord_conf = config.get("landlord", {})

    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        stream = io.StringIO()
        if receipts:
            writer = csv.DictWriter(stream, fieldnames=receipts[0].keys())
            writer.writeheader()
            writer.writerows(receipts)
        zipf.writestr("receipts_data.csv", stream.getvalue())

        for r in receipts:
            tenantName = r.get("Tenant", "Unknown").replace(" ", "_")
            try:
                formatted_date = datetime.datetime.strptime(
                                r.get("Date", ""), "%d %B %Y"
                            ).strftime("%Y%m%d")
            except Exception:
                formatted_date = r.get("Date", "").replace(" ", "")

            custom_filename = f"{tenantName}_{formatted_date}_{r['Bill']}.pdf"
            status = r.get("Status", "ACTIVE")
            folder = "archive" if status == "ARCHIVED" else "active"

            try:
                pdf_stream = generate_professional_pdf(r, landlord_conf)
                zipf.writestr(f"PDFs/{folder}/{custom_filename}", pdf_stream.getvalue())
            except Exception as e:
                print(f"Failed to generate PDF for {r['Bill']}: {e}")

    response = FileResponse(zip_path, media_type="application/zip", filename=zip_filename)
    response.headers["Content-Disposition"] = f'attachment; filename="{zip_filename}"'
    return response

@router.get(Routes.ADMINAPISYNCTEMPLATE, name=Names.DOWNLOADEXCELTEMPLATE)
async def download_excel_template():
    wb = openpyxl.Workbook()
    ws_profile = wb.active
    ws_profile.title = "Tenant_Profile"
    ws_profile.append(PROFILE_HEADERS)
    ws_receipts = wb.create_sheet("Rent_Receipts")
    ws_receipts.append(RECEIPT_HEADERS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for ws in [ws_profile, ws_receipts]:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    # Sample data rows
    ws_profile.append(["T001", "John Doe", "9876543210", "john@gmail.com", "ABC Pvt Ltd", "Delhi", "A101", "MTR001", "", 15000, 500, 8.5, 1000, 300, "Active"])
    ws_profile.append(["T002", "Alice Smith", "9988776655", "alice@gmail.com", "XYZ Ltd", "Noida", "B202", "MTR002", "", 18000, 600, 9.0, 1200, 400, "Active"])
    ws_receipts.append(["T1-001", "T001", "January 2026", "01 Jan 2026", 120, 150, 30, 15000, 500, 255, 1000, 300, 0, 0, 17055, 17055, "PAID", "ACTIVE"])
    ws_receipts.append(["T2-001", "T002", "January 2026", "01 Jan 2026", 80, 110, 30, 18000, 600, 270, 0, 400, 0, 0, 19270, 19270, "PAID", "ACTIVE"])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = StreamingResponse(iter([stream.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = 'attachment; filename="Rent_Data_Template.xlsx"'
    return response

@router.get(Routes.ADMINAPISYNCEXPORTEXCEL, name=Names.EXPORTEXCELDATA)
async def export_excel_data(format: str):
    tenants = load_tenants()
    receipts = get_all_receipts()

    # Build workbook entirely in RAM
    wb = _build_excel_workbook(tenants, receipts)
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    date_str = datetime.datetime.now().strftime('%Y%m%d')

    if format == "xlsx":
        filename = f"Rent_Data_Export_{date_str}.xlsx"
        response = StreamingResponse(
            iter([excel_stream.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    elif format == "zip":
        zip_stream = io.BytesIO()
        with zipfile.ZipFile(zip_stream, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.writestr(f"Rent_Data_Export_{date_str}.xlsx", excel_stream.getvalue())
        zip_stream.seek(0)
        zip_filename = f"Rent_Data_Archive_{date_str}.zip"
        response = StreamingResponse(
            iter([zip_stream.getvalue()]),
            media_type="application/zip"
        )
        response.headers["Content-Disposition"] = f'attachment; filename="{zip_filename}"'
        return response

    elif format == "csv":
        # Convert receipts to CSV
        stream = io.StringIO()
        if receipts:
            writer = csv.DictWriter(stream, fieldnames=receipts[0].keys())
            writer.writeheader()
            writer.writerows(receipts)
        else:
            stream.write("No data found.")
        
        csv_filename = f"receipts_export_{date_str}.csv"
        response = StreamingResponse(
            iter([stream.getvalue()]),
            media_type="text/csv"
        )
        response.headers["Content-Disposition"] = f'attachment; filename="{csv_filename}"'
        return response

    raise HTTPException(status_code=400, detail="Unsupported format. Use 'xlsx', 'zip', or 'csv'.")

def parse_excel_bytes(file_bytes, filename):
    wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    if "Tenant_Profile" not in wb.sheetnames or "Rent_Receipts" not in wb.sheetnames:
        raise ValueError(f"File '{filename}' is missing required sheets 'Tenant_Profile' and/or 'Rent_Receipts'.")

    ws_profile = wb["Tenant_Profile"]
    ws_receipts = wb["Rent_Receipts"]

    p_headers = [str(cell.value).strip() if cell.value else f"Col{i}" for i, cell in enumerate(ws_profile[1])]
    r_headers = [str(cell.value).strip() if cell.value else f"Col{i}" for i, cell in enumerate(ws_receipts[1])]

    tenants_dict = {}
    for row in ws_profile.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        row_dict = {p_headers[i]: (str(val).strip() if val is not None else "") for i, val in enumerate(row)}
        t_id = row_dict.get("tenantId", "")
        if t_id:
            tenants_dict[t_id] = {"profile": row_dict, "receipts": []}

    for row in ws_receipts.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        row_dict = {r_headers[i]: (str(val).strip() if val is not None else "") for i, val in enumerate(row)}
        t_id = row_dict.get("tenantId", "")
        if t_id in tenants_dict:
            tenants_dict[t_id]["receipts"].append(row_dict)

    return tenants_dict


# ============================================================================
# IMPORT VALIDATION & CONFLICT DETECTION HELPERS
# ============================================================================

def _is_encrypted_pin(pin_value: str) -> bool:
    """
    Detect if a PIN value appears to be encrypted rather than a plain 4-digit PIN.
    
    Encrypted PINs typically:
    - Are longer than 4 characters
    - Contain base64 characters (A-Z, a-z, 0-9, +, /, =)
    - Don't match simple 4-digit pattern
    """
    if not pin_value or not str(pin_value).strip():
        return False
    
    pin_str = str(pin_value).strip()
    
    # If it's exactly 4 digits, it's a plain PIN
    if len(pin_str) == 4 and pin_str.isdigit():
        return False
    
    # If it's longer than 20 chars and contains base64-like characters, likely encrypted
    if len(pin_str) > 20:
        import re
        base64_pattern = re.compile(r'^[A-Za-z0-9+/=]+$')
        if base64_pattern.match(pin_str):
            return True
    
    # If it's not 4 digits but has alphanumeric/special chars, likely encrypted
    if len(pin_str) > 4:
        return True
    
    return False


def _generate_random_pin() -> str:
    """Generate a random 4-digit PIN for auto-assignment."""
    import random
    import secrets
    # Use secrets for cryptographically secure random generation
    return f"{secrets.randbelow(10000):04d}"


def _extract_numeric_tenant_id(tenant_id_str: str) -> int:
    """Extract numeric tenant ID from format like 'T001' -> 1."""
    import re
    # Remove 'T' prefix and leading zeros
    cleaned = re.sub(r'^[Tt]', '', str(tenant_id_str).strip())
    try:
        return int(cleaned)
    except ValueError:
        return 0


def _get_next_available_tenant_id() -> int:
    """Get the next available tenant ID number."""
    tenants = load_tenants(include_archived=True)
    if not tenants:
        return 1
    max_id = max(t.id for t in tenants)
    return max_id + 1


def _format_tenant_id(num_id: int) -> str:
    """Format numeric ID as 'T001' style."""
    return f"T{str(num_id).zfill(3)}"


def _remap_bill_no(original_bill_no: str, old_tenant_id_str: str, new_tenant_id: int) -> str:
    """
    Remap a receipt bill number from the old tenant prefix to a new one.
    Example: 'T1-001' with old_tenant_id_str='T001' and new_tenant_id=2 → 'T2-001'
    Falls back to the original value if the pattern doesn't match.
    """
    if not original_bill_no:
        return original_bill_no
    old_numeric = _extract_numeric_tenant_id(old_tenant_id_str)
    old_prefix = f"T{old_numeric}-"
    if original_bill_no.startswith(old_prefix):
        seq_part = original_bill_no[len(old_prefix):]
        return f"T{new_tenant_id}-{seq_part}"
    return original_bill_no


def _detect_import_conflicts(parsed_data: dict) -> dict:
    """
    Detect tenant and receipt conflicts between import data and existing system data.
    
    Returns a dict mapping target_key -> conflict_info for each conflict.
    """
    from app.services.billing_service import get_all_receipts
    sys_tenants = load_tenants(include_archived=True)
    sys_receipts = get_all_receipts()
    
    sys_tenant_ids = {t.id for t in sys_tenants}
    sys_tenant_names = {t.name.lower(): t for t in sys_tenants}
    sys_tenant_phones = {t.phone.lower(): t for t in sys_tenants if t.phone}
    sys_tenant_emails = {t.email.lower(): t for t in sys_tenants if getattr(t, 'email', '')}
    sys_tenant_meters = {t.meterId.lower(): t for t in sys_tenants if getattr(t, 'meterId', '')}
    
    sys_receipt_bills = {r.get("Bill") for r in sys_receipts}
    # Key by tenantId+month (ID-based, rename-safe) for duplicate month detection.
    # Falls back to TenantId from receipt dict; name is NOT used for identity here.
    sys_receipt_tenant_months = {
        f"{int(r.get('TenantId', 0) or 0)}_{r.get('Month', '').lower()}"
        for r in sys_receipts
        if int(r.get('TenantId', 0) or 0) > 0
    }
    
    conflicts = {}
    
    for t_id, t_data in parsed_data.items():
        p = t_data["profile"]
        t_name = p.get("tenantName", "").strip()
        t_phone = p.get("Phone", "").strip()
        t_email = p.get("Email", "").strip()
        t_meter = p.get("meterId", "").strip()
        
        numeric_id = _extract_numeric_tenant_id(t_id)
        
        conflict_info = {
            "importTenant": {
                "tenantId": t_id,
                "tenantName": t_name
            },
            "matches": [],
            "receiptConflicts": []
        }
        
        # 1. Tenant ID match
        if numeric_id in sys_tenant_ids:
            existing = next((t for t in sys_tenants if t.id == numeric_id), None)
            if existing:
                conflict_info["matches"].append({
                    "type": "tenant_id",
                    "existingTenantId": existing.id,
                    "existingTenantName": existing.name
                })
                
        # 2. Name match
        if t_name.lower() in sys_tenant_names:
            existing = sys_tenant_names[t_name.lower()]
            conflict_info["matches"].append({
                "type": "name",
                "existingTenantId": existing.id,
                "existingTenantName": existing.name
            })
            
        # 3. Phone match
        if t_phone and t_phone.lower() in sys_tenant_phones:
            existing = sys_tenant_phones[t_phone.lower()]
            conflict_info["matches"].append({
                "type": "phone",
                "existingTenantId": existing.id,
                "existingTenantName": existing.name
            })
            
        # 4. Email match
        if t_email and t_email.lower() in sys_tenant_emails:
            existing = sys_tenant_emails[t_email.lower()]
            conflict_info["matches"].append({
                "type": "email",
                "existingTenantId": existing.id,
                "existingTenantName": existing.name
            })
            
        # 5. Meter ID match
        if t_meter and t_meter.lower() in sys_tenant_meters:
            existing = sys_tenant_meters[t_meter.lower()]
            conflict_info["matches"].append({
                "type": "meterId",
                "existingTenantId": existing.id,
                "existingTenantName": existing.name
            })
            
        # Remove duplicate matches (same tenant matched multiple ways)
        unique_matches = []
        seen_match_keys = set()
        for m in conflict_info["matches"]:
            key = f"{m['type']}_{m['existingTenantId']}"
            if key not in seen_match_keys:
                seen_match_keys.add(key)
                unique_matches.append(m)
        conflict_info["matches"] = unique_matches

        # 6. Receipt matches
        for r in t_data.get("receipts", []):
            billNo = str(r.get("BillNo", "")).strip()
            month = _parse_month_date(str(r.get("Month", "")).strip())
            
            conflict_reason = None
            if billNo and billNo in sys_receipt_bills:
                conflict_reason = "billNo_exists"
            elif month:
                # Use tenant_id+month duplicate check (ID-based, rename-safe).
                # The numeric_id here is the imported file's tenant ID; during
                # execute it will be remapped to the resolved existingTenantId.
                if f"{numeric_id}_{month.lower()}" in sys_receipt_tenant_months:
                    conflict_reason = "tenant_month_exists"
                
            if conflict_reason:
                conflict_info["receiptConflicts"].append({
                    "billNo": billNo,
                    "month": month,
                    "reason": conflict_reason,
                    "actionRequired": True
                })
                
        if conflict_info["matches"] or conflict_info["receiptConflicts"]:
            conflicts[f"{t_id}"] = conflict_info
            
    return conflicts


def _detect_encrypted_pins(parsed_data: dict) -> dict:
    """
    Detect encrypted PINs in the import data.
    
    Returns a dict mapping target_key -> pin_info for each encrypted PIN.
    """
    encrypted_pins = {}
    
    for t_id, t_data in parsed_data.items():
        p = t_data["profile"]
        pin_value = str(p.get("PIN") or "").strip()
        
        if pin_value and _is_encrypted_pin(pin_value):
            encrypted_pins[t_id] = {
                "tenantId": t_id,
                "tenantName": p.get("tenantName", ""),
                "pin_value": pin_value[:20] + "..." if len(pin_value) > 20 else pin_value,
                "pin_length": len(pin_value),
                "is_encrypted": True
            }
    
    return encrypted_pins


# ============================================================================
# ENHANCED PREVIEW ENDPOINT (with conflict & encrypted PIN detection)
# ============================================================================

@router.post(Routes.ADMINAPISYNCIMPORTPREVIEW, name=Names.IMPORTPREVIEWDATA)
async def import_preview_data(files: List[UploadFile] = File(...)):
    preview_data = {}
    try:
        for file in files:
            content = await file.read()
            if file.filename.endswith('.zip'):
                with zipfile.ZipFile(io.BytesIO(content)) as z:
                    for zip_info in z.infolist():
                        if zip_info.filename.endswith('.xlsx'):
                            with z.open(zip_info) as f:
                                preview_data[zip_info.filename] = parse_excel_bytes(f.read(), zip_info.filename)
            elif file.filename.endswith('.xlsx'):
                preview_data[file.filename] = parse_excel_bytes(content, file.filename)
            else:
                raise HTTPException(status_code=400, detail="Only .xlsx or .zip files are supported.")
        
        # ── CONFLICT DETECTION ──
        all_conflicts = {}
        all_encrypted_pins = {}
        
        for filename, parsed_data in preview_data.items():
            # Detect tenant ID conflicts
            conflicts = _detect_import_conflicts(parsed_data)
            if conflicts:
                all_conflicts[filename] = conflicts
            
            # Detect encrypted PINs
            encrypted_pins = _detect_encrypted_pins(parsed_data)
            if encrypted_pins:
                all_encrypted_pins[filename] = encrypted_pins
        
        return {
            "status": "success",
            "files": preview_data,
            "conflicts": all_conflicts if all_conflicts else {},
            "encrypted_pins": all_encrypted_pins if all_encrypted_pins else {},
            "requires_resolution": bool(all_conflicts or all_encrypted_pins),
            "predicted_next_tenant_id": _get_next_available_tenant_id()
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


def _parse_excel_date(val: str) -> str:
    """Parse Excel date strings/serials and format as 'dd MMMM yyyy'."""
    if not val or not str(val).strip():
        return ""
    val_str = str(val).strip()

    # Try ISO format: 2026-06-01 00:00:00 or 2026-06-01T00:00:00
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            dt = datetime.datetime.strptime(val_str, fmt)
            return dt.strftime("%d %B %Y")  # e.g., "01 June 2026"
        except ValueError:
            continue

    # Try Excel serial number
    try:
        serial = float(val_str)
        # Excel epoch is 1899-12-30 (with the infamous 1900 leap year bug)
        excel_epoch = datetime.datetime(1899, 12, 30)
        dt = excel_epoch + datetime.timedelta(days=serial)
        return dt.strftime("%d %B %Y")
    except (ValueError, OverflowError):
        pass

    # Return original if we can't parse
    return val_str


def _parse_month_date(val: str) -> str:
    """Parse Excel month value to 'Month Year' format."""
    if not val or not str(val).strip():
        return ""
    val_str = str(val).strip()

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            dt = datetime.datetime.strptime(val_str, fmt)
            return dt.strftime("%B %Y")  # e.g., "June 2026"
        except ValueError:
            continue

    # Try Excel serial
    try:
        serial = float(val_str)
        excel_epoch = datetime.datetime(1899, 12, 30)
        dt = excel_epoch + datetime.timedelta(days=serial)
        return dt.strftime("%B %Y")
    except (ValueError, OverflowError):
        pass

    return val_str

from typing import Dict, Literal

VALID_TENANT_STATUSES = {"Active", "Inactive", "Archived"}

def normalize_tenant_status(value: str | None, default: str = "Active") -> str:
    candidate = str(value or "").strip().title()
    return candidate if candidate in VALID_TENANT_STATUSES else default


# ============================================================================
# ENHANCED EXECUTE IMPORT ENDPOINT
# ============================================================================

@router.post(Routes.ADMINAPISYNCIMPORTEXECUTE, name=Names.IMPORTEXECUTEDATA)
async def import_execute_data(
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...),
    selectedtargets: Optional[str] = Form(None),
    selectedTargets: Optional[str] = Form(None),
    targetstatuses: Optional[str] = Form(None),
    # ── NEW: Conflict resolution parameters ──
    idresolutions: Optional[str] = Form(None),
    pinhandling: Optional[str] = Form("prompt"),
    pinresolutions: Optional[str] = Form(None),
    receiptstrategies: Optional[str] = Form(None), # JSON: { "filename::t_id": "SKIP" | "MERGE_RECEIPTS_ONLY" | "REPLACE_RECEIPTS", ... }
):
    """
    Execute import with conflict resolution support inside a single transaction.
    """
    targets = selectedtargets or selectedTargets or ""
    if not targets:
        raise HTTPException(status_code=400, detail="selectedtargets is required")
    
    try:
        selected_list = json.loads(targets)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON in selectedtargets")

    if not isinstance(selected_list, list):
        raise HTTPException(status_code=400, detail="selectedtargets must be a JSON array")

    try:
        status_overrides: Dict[str, str] = json.loads(targetstatuses or "{}")
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON in targetstatuses.")

    id_resolutions: Dict[str, str] = {}
    if idresolutions:
        try:
            id_resolutions = json.loads(idresolutions)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="Invalid JSON in idresolutions.")
    
    pin_resolutions: Dict[str, str] = {}
    if pinresolutions:
        try:
            pin_resolutions = json.loads(pinresolutions)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="Invalid JSON in pinresolutions.")
            
    receipt_strategies: Dict[str, str] = {}
    if receiptstrategies:
        try:
            receipt_strategies = json.loads(receiptstrategies)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="Invalid JSON in receiptstrategies.")

    pin_handling_mode = (pinhandling or "prompt").strip().lower()
    if pin_handling_mode not in {"prompt", "skip", "assign_random"}:
        raise HTTPException(status_code=400, detail="Invalid pinhandling. Use 'prompt', 'skip', or 'assign_random'.")

    if not selected_list:
        raise HTTPException(status_code=400, detail="No tenants selected for import.")

    # Parse all files in memory first
    parsed_files_data = {}
    temp_files_to_cleanup = []
    
    try:
        for file in files:
            content = await file.read()
            temp_files_to_cleanup.append(file)
            
            if file.filename.endswith(".zip"):
                with zipfile.ZipFile(io.BytesIO(content)) as z:
                    for zip_info in z.infolist():
                        if zip_info.filename.endswith(".xlsx"):
                            with z.open(zip_info) as f:
                                parsed_files_data[zip_info.filename] = parse_excel_bytes(f.read(), zip_info.filename)
            elif file.filename.endswith(".xlsx"):
                parsed_files_data[file.filename] = parse_excel_bytes(content, file.filename)
    except Exception as e:
        for tf in temp_files_to_cleanup:
            try: await tf.close()
            except: pass
        raise HTTPException(status_code=400, detail=f"Failed to parse files: {str(e)}")

    sys_tenants = load_tenants(include_archived=True)
    sys_tenant_ids = {t.id for t in sys_tenants}
    # ID-indexed for validation; NOT used for name-based ownership resolution.
    sys_tenant_by_id = {t.id: t for t in sys_tenants}

    # Detect conflicts across all files
    unresolved_conflicts = []
    unresolved_pins = []

    for filename, parsed_data in parsed_files_data.items():
        conflicts = _detect_import_conflicts(parsed_data)
        encrypted_pins = _detect_encrypted_pins(parsed_data)
        
        for t_id, conflict_info in conflicts.items():
            target_key = f"{filename}::{t_id}"
            if target_key not in selected_list:
                continue
                
            has_tenant_conflict = len(conflict_info.get("matches", [])) > 0
            has_receipt_conflict = len(conflict_info.get("receiptConflicts", [])) > 0
            
            resolution_action = id_resolutions.get(target_key)
            if has_tenant_conflict and resolution_action not in ("CREATE_NEW", "UPDATE_EXISTING", "SKIP", "MERGE_RECEIPTS_ONLY"):
                unresolved_conflicts.append({
                    "target": target_key,
                    "tenantName": conflict_info["importTenant"]["tenantName"],
                    "reason": "unresolved_tenant_conflict",
                    "matches": conflict_info.get("matches")
                })
                continue
                
            receipt_action = receipt_strategies.get(target_key)
            if has_receipt_conflict and receipt_action not in ("SKIP", "MERGE_RECEIPTS_ONLY", "REPLACE_RECEIPTS") and resolution_action != "SKIP":
                unresolved_conflicts.append({
                    "target": target_key,
                    "tenantName": conflict_info["importTenant"]["tenantName"],
                    "reason": "unresolved_receipt_conflict",
                    "receiptConflicts": conflict_info.get("receiptConflicts")
                })
        
        if pin_handling_mode == "prompt":
            for t_id, pin_info in encrypted_pins.items():
                target_key = f"{filename}::{t_id}"
                if target_key not in selected_list:
                    continue
                if target_key not in pin_resolutions and id_resolutions.get(target_key) != "SKIP":
                    unresolved_pins.append({
                        "target": target_key,
                        "tenantName": pin_info["tenantName"],
                        "reason": "encrypted_pin_detected"
                    })
                    
    if unresolved_conflicts or unresolved_pins:
        for tf in temp_files_to_cleanup:
            try: await tf.close()
            except: pass
        raise HTTPException(
            status_code=409,
            detail={
                "message": "Import requires manual resolution.",
                "conflicts": unresolved_conflicts,
                "encrypted_pins": unresolved_pins
            }
        )

    # Pre-compute a target_key → existing_tenant_id map from conflict detector
    # results (which already resolved tenants by ID, not by name). This map is
    # the authoritative source for UPDATE_EXISTING and MERGE_RECEIPTS_ONLY.
    existing_tenant_id_map: dict[str, int] = {}
    for filename, parsed_data in parsed_files_data.items():
        conflicts = _detect_import_conflicts(parsed_data)
        for t_id, conflict_info in conflicts.items():
            target_key = f"{filename}::{t_id}"
            matches = conflict_info.get("matches", [])
            # Pick the first match (the ID-match, if present, is always first).
            if matches:
                existing_tenant_id_map[target_key] = matches[0]["existingTenantId"]

    # Schedule backup BEFORE processing
    background_tasks.add_task(create_full_backup, tag="pre_import_excel")

    imported_tenants = []
    imported_receipts = 0
    skipped_targets = set(selected_list)
    auto_assigned_pins = {}
    
    admin_username = "Admin"
    job_result = {"items": []}

    try:
        with get_conn() as conn:
            # Start transaction explicitly
            conn.execute("BEGIN")
            
            # Create import job record
            job_cur = conn.execute(
                "INSERT INTO import_jobs (created_at, created_by, filename, status, preview_json, resolution_json, result_json) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (datetime.datetime.utcnow().isoformat(), admin_username, ", ".join(parsed_files_data.keys()), "IN_PROGRESS", "{}", "{}", "{}")
            )
            job_id = job_cur.lastrowid
            
            for filename, parsed_data in parsed_files_data.items():
                for t_id, t_data in parsed_data.items():
                    target_key = f"{filename}::{t_id}"
                    if target_key not in selected_list:
                        continue
                        
                    skipped_targets.discard(target_key)
                    action = id_resolutions.get(target_key, "CREATE_NEW")
                    
                    if action == "SKIP":
                        conn.execute(
                            "INSERT INTO import_job_items (import_job_id, target_key, import_tenant_id, import_tenant_name, action, result) VALUES (?, ?, ?, ?, ?, ?)",
                            (job_id, target_key, t_id, t_data["profile"].get("tenantName", ""), action, "SKIPPED")
                        )
                        continue

                    p = t_data["profile"]
                    t_name = p.get("tenantName", "").strip()
                    if not t_name:
                        continue

                    # Resolve existing tenant by ID from pre-computed conflict map
                    # (never by name). CREATE_NEW does not need this.
                    existing_tid = existing_tenant_id_map.get(target_key)
                    existing_t = sys_tenant_by_id.get(existing_tid) if existing_tid else None

                    tenantId = None
                    is_new = False

                    if action == "CREATE_NEW":
                        # Insert new tenant
                        next_id = _get_next_available_tenant_id()
                        while next_id in sys_tenant_ids:
                            next_id += 1
                        tenantId = next_id

                        viewToken = str(uuid.uuid4())
                        conn.execute('''
                            INSERT INTO tenants (
                                id, name, company, phone, email, address, roomnumber, occupation, notes, status,
                                rent, water, electricityrate, previousmeter, additionalpersoncharge, securitydeposit,
                                defaulttankWatercharge, meterid, viewToken, tenantpin, failed_attempts
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            tenantId, t_name, p.get("Company", ""), p.get("Phone", ""), p.get("Email", ""),
                            p.get("Address", ""), p.get("Room", ""), "", "", normalize_tenant_status(status_overrides.get(target_key), p.get("Status", "Active")),
                            float(p.get("Rent", 0) or 0), float(p.get("Water", 0) or 0), float(p.get("electricityRate", 0) or 0),
                            0, float(p.get("additionalPersonRate", 0) or 0), 0, float(p.get("tankWater", 0) or 0),
                            p.get("meterId", ""), viewToken, "", 0
                        ))
                        sys_tenant_ids.add(tenantId)
                        is_new = True

                    elif action == "UPDATE_EXISTING" and existing_t:
                        tenantId = existing_t.id
                        conn.execute('''
                            UPDATE tenants SET
                                company=COALESCE(?, company), phone=COALESCE(?, phone), email=COALESCE(?, email),
                                address=COALESCE(?, address), roomnumber=COALESCE(?, roomnumber), meterid=COALESCE(?, meterid),
                                rent=COALESCE(?, rent), water=COALESCE(?, water), electricityrate=COALESCE(?, electricityrate),
                                additionalpersoncharge=COALESCE(?, additionalpersoncharge), defaulttankWatercharge=COALESCE(?, defaulttankWatercharge),
                                status=COALESCE(?, status)
                            WHERE id=?
                        ''', (
                            p.get("Company"), p.get("Phone"), p.get("Email"), p.get("Address"), p.get("Room"), p.get("meterId"),
                            float(p.get("Rent", 0) or 0) if p.get("Rent") else None,
                            float(p.get("Water", 0) or 0) if p.get("Water") else None,
                            float(p.get("electricityRate", 0) or 0) if p.get("electricityRate") else None,
                            float(p.get("additionalPersonRate", 0) or 0) if p.get("additionalPersonRate") else None,
                            float(p.get("tankWater", 0) or 0) if p.get("tankWater") else None,
                            normalize_tenant_status(status_overrides.get(target_key), p.get("Status", "Active")),
                            tenantId
                        ))

                    elif action == "MERGE_RECEIPTS_ONLY" and existing_t:
                        tenantId = existing_t.id
                        # Don't update tenant profile

                    if not tenantId:
                        continue  # Should not happen based on validation

                    # ── PIN HANDLING ──
                    if action in ("CREATE_NEW", "UPDATE_EXISTING"):
                        raw_pin = str(p.get("PIN") or "").strip()
                        plain_pin = None
                        pin_changed = False
                        hashed_pin = None
                        encrypted_pin = None
                        
                        if raw_pin:
                            if _is_encrypted_pin(raw_pin):
                                if pin_handling_mode == "assign_random":
                                    plain_pin = _generate_random_pin()
                                    auto_assigned_pins[target_key] = plain_pin
                                    pin_changed = True
                                elif pin_handling_mode == "prompt":
                                    plain_pin = pin_resolutions.get(target_key)
                                    if plain_pin:
                                        pin_changed = True
                            else:
                                plain_pin = raw_pin
                                pin_changed = True
                                
                        if pin_changed and plain_pin:
                            try:
                                validate_tenantPin(plain_pin)
                                hashed_pin = hash_pin(plain_pin)
                                encrypted_pin = encrypt_admin_view_pin(plain_pin)
                                
                                conn.execute("UPDATE tenants SET tenantpin = ? WHERE id = ?", (hashed_pin, tenantId))
                                now_iso = datetime.datetime.utcnow().isoformat()
                                conn.execute("INSERT INTO tenantPin_history (tenantId, pin_hash, changed_at) VALUES (?, ?, ?)", (tenantId, hashed_pin, now_iso))
                                conn.execute("INSERT OR REPLACE INTO tenantPin_admin_store (tenantId, encrypted_pin, updated_at) VALUES (?, ?, ?)", (tenantId, encrypted_pin, now_iso))
                                if not is_new:
                                    conn.execute("DELETE FROM tenant_sessions WHERE tenantId = ?", (tenantId,))
                            except HTTPException:
                                pass # Invalid pin format
                                
                    # ── RECEIPTS ──
                    rec_strategy = receipt_strategies.get(target_key, "MERGE_RECEIPTS_ONLY")
                    if rec_strategy == "REPLACE_RECEIPTS":
                        conn.execute("DELETE FROM receipts WHERE tenantId = ?", (tenantId,))
                        
                    if rec_strategy in ("MERGE_RECEIPTS_ONLY", "REPLACE_RECEIPTS"):
                        for r in t_data.get("receipts", []):
                            original_billNo = r.get("BillNo", "").strip()
                            if not original_billNo: continue
                            
                            # Remap bill number prefix when tenant was created as new
                            if action == "CREATE_NEW":
                                billNo = _remap_bill_no(original_billNo, t_id, tenantId)
                            else:
                                billNo = original_billNo
                            
                            r_date = _parse_excel_date(r.get("Date", ""))
                            r_month = _parse_month_date(r.get("Month", ""))
                            
                            exists = conn.execute("SELECT 1 FROM receipts WHERE billNo = ?", (billNo,)).fetchone()
                            
                            if exists:
                                if rec_strategy == "MERGE_RECEIPTS_ONLY":
                                    conn.execute("""
                                        UPDATE receipts SET
                                            date=?, month=?, tenantId=?, tenant=?, previous=?, current=?, units=?, rent=?,
                                            additional=?, water=?, tankWater=?, electricity=?, total=?, pdf=?,
                                            rate=?, status=?, additionalpersonrate=?,
                                            paymentstatus=?, maintenancecharge=?, maintenancedesc=?, previousarrears=?, amountreceived=?
                                        WHERE billNo=?
                                    """, (
                                        r_date, r_month, tenantId, t_name, float(r.get("Previous", 0) or 0), float(r.get("Current", 0) or 0),
                                        float(r.get("Units", 0) or 0), float(r.get("Rent", 0) or 0), float(r.get("Additional", 0) or 0), 
                                        float(r.get("Water", 0) or 0), float(r.get("tankWater", 0) or 0), float(r.get("Electricity", 0) or 0), 
                                        float(r.get("Total", 0) or 0), "", float(r.get("Rate", 0) or 0), r.get("receiptStatus", "ACTIVE"), 
                                        float(r.get("additionalPersonRate", 0) or 0), r.get("paymentStatus", "PENDING"), 
                                        float(r.get("Maintenance", 0) or 0), r.get("MaintenanceDesc", ""), float(r.get("Arrears", 0) or 0), 
                                        float(r.get("amountReceived", 0) or 0), billNo
                                    ))
                                    imported_receipts += 1
                            else:
                                conn.execute("""
                                    INSERT INTO receipts (
                                        billNo, date, month, tenantId, tenant, previous, current, units, rent,
                                        additional, water, tankWater, electricity, total, pdf,
                                        tenantphone, tenantcompany, tenantaddress, rate, status,
                                        additionalpersons, additionalpersonrate, receiptversion, generatedby, paymentstatus,
                                        maintenancecharge, maintenancedesc, previousarrears, amountreceived
                                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """, (
                                    billNo, r_date, r_month, tenantId, t_name, float(r.get("Previous", 0) or 0), float(r.get("Current", 0) or 0),
                                    float(r.get("Units", 0) or 0), float(r.get("Rent", 0) or 0), float(r.get("Additional", 0) or 0), 
                                    float(r.get("Water", 0) or 0), float(r.get("tankWater", 0) or 0), float(r.get("Electricity", 0) or 0), 
                                    float(r.get("Total", 0) or 0), "", "", "", "", float(r.get("Rate", 0) or 0), r.get("receiptStatus", "ACTIVE"),
                                    0, float(r.get("additionalPersonRate", 0) or 0), 8, "Import", r.get("paymentStatus", "PENDING"),
                                    float(r.get("Maintenance", 0) or 0), r.get("MaintenanceDesc", ""), float(r.get("Arrears", 0) or 0), float(r.get("amountReceived", 0) or 0)
                                ))
                                imported_receipts += 1
                                
                    conn.execute(
                        "INSERT INTO import_job_items (import_job_id, target_key, import_tenant_id, import_tenant_name, action, existing_tenant_id, result) VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (job_id, target_key, t_id, t_name, action, existing_t.id if existing_t else None, "SUCCESS")
                    )
                    
                    imported_tenants.append({
                        "target": target_key,
                        "tenantId": tenantId,
                        "tenantName": t_name,
                        "action": action
                    })

            # Mark job complete
            conn.execute("UPDATE import_jobs SET status = ?, result_json = ? WHERE id = ?", ("COMPLETED", json.dumps({"tenants": len(imported_tenants), "receipts": imported_receipts}), job_id))
            
            # Commit the single transaction
            conn.commit()

        msg_parts = [f"Import completed successfully."]
        msg_parts.append(f"Tenants: {len(imported_tenants)} processed.")
        msg_parts.append(f"Receipts: {imported_receipts} imported/updated.")
        
        if auto_assigned_pins and pin_handling_mode == "assign_random":
            msg_parts.append(f"Auto-assigned PINs for {len(auto_assigned_pins)} tenant(s).")
        if skipped_targets:
            msg_parts.append(f"Warning: {len(skipped_targets)} selected target(s) not found in files.")

        response_data = {
            "status": "success",
            "message": " ".join(msg_parts),
            "tenants": len(imported_tenants),
            "receipts": imported_receipts,
            "imported_tenants": imported_tenants,
            "unmatched_targets": list(skipped_targets) if skipped_targets else []
        }
        if auto_assigned_pins:
            response_data["auto_assigned_pins"] = [{"target": k, "pin": v} for k, v in auto_assigned_pins.items()]
        
        return response_data
        
    except Exception as e:
        # Implicit rollback occurs if an exception is raised within the context manager before commit()
        raise HTTPException(status_code=400, detail=f"Import execution failed: {str(e)}")
        
    finally:
        for tf in temp_files_to_cleanup:
            try: await tf.close()
            except: pass
            try:
                temp_path = getattr(tf.file, "name", None)
                if isinstance(temp_path, str) and os.path.isfile(temp_path):
                    os.remove(temp_path)
            except: pass


@router.get(Routes.ADMINAPIBILLINGARCHIVEDATA)
async def get_archive_data():
    tenants = load_tenants(include_archived=True)
    archivedtenants = [
        tenant for tenant in tenants
        if (getattr(tenant, "status", "") or "").strip().lower() == "archived"
    ]
    archivedtenantids = {int(tenant.id) for tenant in archivedtenants}

    receipts = get_all_receipts(include_archived_tenants=True)
    archivedreceipts = [
        receipt for receipt in receipts
        if str(receipt.get("Status", "") or "").strip().upper() == "ARCHIVED"
        or int(receipt.get("TenantId", 0) or 0) in archivedtenantids
    ]

    archivedreceipts.sort(
        key=lambda r: (
            int(r.get("TenantId", 0) or 0),
            str(r.get("Date", "") or ""),
            str(r.get("Bill", "") or ""),
        ),
        reverse=True,
    )

    return {
        "tenants": archivedtenants,
        "receipts": archivedreceipts,
    }

if __name__ == "__main__":
    sys_conf = config.get("system", {})
    server_host = sys_conf["server"]["host"]
    server_port = sys_conf["server"]["port"]
    is_debug = sys_conf["server"]["debug"]

    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('10.255.255.255', 1))
        local_ip = s.getsockname()[0]
    except Exception:
        local_ip = '127.0.0.1'
    finally:
        s.close()

    print(f"\\n{'='*50}")
    print(f" {sys_conf['app']['title']} is starting...")
    print(f"{'='*50}")
    print(f" [Local]:   http://127.0.0.1:{server_port}")
    print(f" [Network]: http://{local_ip}:{server_port}")
    print(f" [Note]:    Do NOT click the {server_host} link below")
    print(f"{'='*50}\\n")

    uvicorn.run(
        "app:app",
        host=server_host,
        port=server_port,
        reload=is_debug,
        proxy_headers=True,
        forwarded_allow_ips="*",
        access_log=True
    )

# # File: app/app/api/sync.py
# from typing import Optional
# from fastapi import APIRouter, HTTPException, UploadFile, File, Form, BackgroundTasks

# from app.core.routes_manifest import Names, Routes

# from fastapi.responses import StreamingResponse, FileResponse
# from app.core.dependencies import config
# from app.models.tenant import Tenant
# import os
# import io
# import json
# import datetime
# import zipfile
# import csv
# from typing import List
# import uvicorn
# import socket

# from app.services.tenant_service import load_tenants, add_tenant, update_tenant
# from app.services.billing_service import get_all_receipts
# from app.services.backup_service import create_full_backup
# from app.core.paths import BACKUPS_DIR

# from app.authentication.common.utils import validate_tenantPin, hash_pin
# from app.authentication.common.pin_vault import encrypt_admin_view_pin
# from app.authentication.tenant.sessions import revoke_all_tenant_sessions
# from app.core.db import get_conn

# import openpyxl
# from openpyxl.styles import Font, PatternFill
# router = APIRouter()


# # ==========================================
# # EXCEL IMPORT & EXPORT ENGINE (RELATIONAL)
# # ==========================================

# PROFILE_HEADERS = [
#     "tenantId", "tenantName", "Phone", "Email", "Company", "Address", "Room",
#     "meterId", "PIN", "Rent", "Water", "electricityRate", "additionalPersonRate",
#     "tankWater", "Status"
# ]

# RECEIPT_HEADERS = [
#     "BillNo", "tenantId", "Month", "Date", "Previous", "Current", "Units", "Rent",
#     "Water", "Electricity", "Additional", "tankWater", "Maintenance", "Arrears",
#     "amountReceived", "Total", "paymentStatus", "receiptStatus"
# ]

# def _build_excel_workbook(tenants_list, receipts_list):
#     """Shared helper: builds and returns an openpyxl workbook in memory."""
#     from app.authentication.common.pin_vault import decrypt_admin_view_pin
#     from app.core.db import get_conn

#     # Pre-fetch all decrypted PINs from admin store
#     decrypted_pins = {}
#     try:
#         with get_conn() as conn:
#             rows = conn.execute("SELECT tenantId, encrypted_pin FROM tenantPin_admin_store").fetchall()
#             for row in rows:
#                 try:
#                     decrypted_pins[row["tenantId"]] = decrypt_admin_view_pin(row["encrypted_pin"])
#                 except Exception:
#                     decrypted_pins[row["tenantId"]] = ""
#     except Exception:
#         pass

#     wb = openpyxl.Workbook()
#     ws_profile = wb.active
#     ws_profile.title = "Tenant_Profile"
#     ws_profile.append(PROFILE_HEADERS)

#     ws_receipts = wb.create_sheet("Rent_Receipts")
#     ws_receipts.append(RECEIPT_HEADERS)

#     header_font = Font(bold=True, color="FFFFFF")
#     header_fill = PatternFill("solid", fgColor="4F81BD")
#     for ws in [ws_profile, ws_receipts]:
#         for cell in ws[1]:
#             cell.font = header_font
#             cell.fill = header_fill

#     tenantId_map = {}
#     for t in tenants_list:
#         t_id_str = f"T{str(t.id).zfill(3)}"
#         tenantId_map[t.name] = t_id_str
#         # Use decrypted PIN for export, fallback to empty string if not available
#         plain_pin = decrypted_pins.get(t.id, "")
#         ws_profile.append([
#             t_id_str, t.name, str(t.phone), getattr(t, 'email', ''), getattr(t, 'company', ''),
#             getattr(t, 'address', ''), getattr(t, 'roomNumber', ''), getattr(t, 'meterId', ''),
#             plain_pin, float(t.rent), float(t.water), float(t.electricityRate),
#             float(t.additionalPersonCharge), float(getattr(t, 'defaulttankWaterCharge', 0.0)), t.status
#         ])

#     for r in receipts_list:
#         t_name = r.get("Tenant", "")
#         t_id_str = tenantId_map.get(t_name, "UNKNOWN")
#         ws_receipts.append([
#             r.get("Bill", ""), t_id_str, r.get("Month", ""), r.get("Date", ""),
#             float(r.get("Previous", 0)), float(r.get("Current", 0)), float(r.get("Units", 0)),
#             float(r.get("Rent", 0)), float(r.get("Water", 0)), float(r.get("Electricity", 0)),
#             float(r.get("Additional", 0)), float(r.get("tankWater", 0)),
#             float(r.get("MaintenanceCharge", 0)), float(r.get("previousArrears", 0)),
#             float(r.get("amountReceived", 0)), float(r.get("Total", 0)),
#             r.get("paymentStatus", "PENDING"), r.get("Status", "ACTIVE")
#         ])

#     return wb


# @router.get(Routes.ADMINAPISYNCEXPORTCSV, name=Names.EXPORTRECEIPTSCSV)
# async def export_receipts_csv(tenants_list: str = "all"):
#     tenants = load_tenants()
#     receipts = get_all_receipts()

#     if tenants_list != "all":
#         selected_ids = [int(x) for x in tenants_list.split(",") if x.isdigit()]
#         selected_names = [t.name for t in tenants if t.id in selected_ids]
#         receipts = [r for r in receipts if r.get("Tenant") in selected_names]

#     stream = io.StringIO()
#     if receipts:
#         writer = csv.DictWriter(stream, fieldnames=receipts[0].keys())
#         writer.writeheader()
#         writer.writerows(receipts)
#     else:
#         stream.write("No data found for selected tenants.")

#     date_str = datetime.datetime.now().strftime('%Y%m%d')
#     filename = f"receipts_export_{date_str}.csv"

#     response = StreamingResponse(iter([stream.getvalue()]), media_type="text/csv")
#     response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
#     return response

# @router.get(Routes.ADMINAPISYNCEXPORTZIP, name=Names.EXPORTFULLZIP)
# async def export_full_zip(tenants_list: str = "all"):
#     tenants = load_tenants()
#     receipts = get_all_receipts()

#     if tenants_list != "all":
#         selected_ids = [int(x) for x in tenants_list.split(",") if x.isdigit()]
#         selected_names = [t.name for t in tenants if t.id in selected_ids]
#         receipts = [r for r in receipts if r.get("Tenant") in selected_names]

#     date_str = datetime.datetime.now().strftime('%Y%m%d')
#     zip_filename = f"tenant_data_{date_str}.zip"
#     zip_path = os.path.join(BACKUPS_DIR, zip_filename)
#     os.makedirs(BACKUPS_DIR, exist_ok=True)

#     from app.services.pdf_service import generate_professional_pdf
#     from app.core.config_service import config
#     landlord_conf = config.get("landlord", {})

#     with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
#         stream = io.StringIO()
#         if receipts:
#             writer = csv.DictWriter(stream, fieldnames=receipts[0].keys())
#             writer.writeheader()
#             writer.writerows(receipts)
#         zipf.writestr("receipts_data.csv", stream.getvalue())

#         for r in receipts:
#             tenantName = r.get("Tenant", "Unknown").replace(" ", "_")
#             try:
#                 formatted_date = datetime.datetime.strptime(
#                                 r.get("Date", ""), "%d %B %Y"
#                             ).strftime("%Y%m%d")
#             except Exception:
#                 formatted_date = r.get("Date", "").replace(" ", "")

#             custom_filename = f"{tenantName}_{formatted_date}_{r['Bill']}.pdf"
#             status = r.get("Status", "ACTIVE")
#             folder = "archive" if status == "ARCHIVED" else "active"

#             try:
#                 pdf_stream = generate_professional_pdf(r, landlord_conf)
#                 zipf.writestr(f"PDFs/{folder}/{custom_filename}", pdf_stream.getvalue())
#             except Exception as e:
#                 print(f"Failed to generate PDF for {r['Bill']}: {e}")

#     response = FileResponse(zip_path, media_type="application/zip", filename=zip_filename)
#     response.headers["Content-Disposition"] = f'attachment; filename="{zip_filename}"'
#     return response

# @router.get(Routes.ADMINAPISYNCTEMPLATE, name=Names.DOWNLOADEXCELTEMPLATE)
# async def download_excel_template():
#     wb = openpyxl.Workbook()
#     ws_profile = wb.active
#     ws_profile.title = "Tenant_Profile"
#     ws_profile.append(PROFILE_HEADERS)
#     ws_receipts = wb.create_sheet("Rent_Receipts")
#     ws_receipts.append(RECEIPT_HEADERS)

#     header_font = Font(bold=True, color="FFFFFF")
#     header_fill = PatternFill("solid", fgColor="4F81BD")
#     for ws in [ws_profile, ws_receipts]:
#         for cell in ws[1]:
#             cell.font = header_font
#             cell.fill = header_fill

#     # Sample data rows
#     ws_profile.append(["T001", "John Doe", "9876543210", "john@gmail.com", "ABC Pvt Ltd", "Delhi", "A101", "MTR001", "", 15000, 500, 8.5, 1000, 300, "Active"])
#     ws_profile.append(["T002", "Alice Smith", "9988776655", "alice@gmail.com", "XYZ Ltd", "Noida", "B202", "MTR002", "", 18000, 600, 9.0, 1200, 400, "Active"])
#     ws_receipts.append(["T1-001", "T001", "January 2026", "01 Jan 2026", 120, 150, 30, 15000, 500, 255, 1000, 300, 0, 0, 17055, 17055, "PAID", "ACTIVE"])
#     ws_receipts.append(["T2-001", "T002", "January 2026", "01 Jan 2026", 80, 110, 30, 18000, 600, 270, 0, 400, 0, 0, 19270, 19270, "PAID", "ACTIVE"])

#     stream = io.BytesIO()
#     wb.save(stream)
#     stream.seek(0)
#     response = StreamingResponse(iter([stream.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response.headers["Content-Disposition"] = 'attachment; filename="Rent_Data_Template.xlsx"'
#     return response

# # @router.get(Routes.ADMINAPISYNCEXPORTEXCEL, name=Names.EXPORTEXCELDATA)
# # async def export_excel_data(format: str):
# #     tenants = load_tenants()
# #     receipts = get_all_receipts()

# #     # Build workbook entirely in RAM
# #     wb = _build_excel_workbook(tenants, receipts)
# #     excel_stream = io.BytesIO()
# #     wb.save(excel_stream)
# #     excel_stream.seek(0)

# #     date_str = datetime.datetime.now().strftime('%Y%m%d')

# #     if format == "xlsx":
# #         filename = f"Rent_Data_Export_{date_str}.xlsx"
# #         response = StreamingResponse(
# #             iter([excel_stream.getvalue()]),
# #             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
# #         )
# #         response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
# #         return response

# #     elif format == "zip":
# #         zip_stream = io.BytesIO()
# #         with zipfile.ZipFile(zip_stream, 'w', zipfile.ZIP_DEFLATED) as zipf:
# #             zipf.writestr(f"Rent_Data_Export_{date_str}.xlsx", excel_stream.getvalue())
# #         zip_stream.seek(0)
# #         zip_filename = f"Rent_Data_Archive_{date_str}.zip"
# #         response = StreamingResponse(
# #             iter([zip_stream.getvalue()]),
# #             media_type="application/zip"
# #         )
# #         response.headers["Content-Disposition"] = f'attachment; filename="{zip_filename}"'
# #         return response

# #     raise HTTPException(status_code=400, detail="Unsupported format. Use 'xlsx' or 'zip'.")
# @router.get(Routes.ADMINAPISYNCEXPORTEXCEL, name=Names.EXPORTEXCELDATA)
# async def export_excel_data(format: str):
#     tenants = load_tenants()
#     receipts = get_all_receipts()

#     # Build workbook entirely in RAM
#     wb = _build_excel_workbook(tenants, receipts)
#     excel_stream = io.BytesIO()
#     wb.save(excel_stream)
#     excel_stream.seek(0)

#     date_str = datetime.datetime.now().strftime('%Y%m%d')

#     if format == "xlsx":
#         filename = f"Rent_Data_Export_{date_str}.xlsx"
#         response = StreamingResponse(
#             iter([excel_stream.getvalue()]),
#             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )
#         response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
#         return response

#     elif format == "zip":
#         zip_stream = io.BytesIO()
#         with zipfile.ZipFile(zip_stream, 'w', zipfile.ZIP_DEFLATED) as zipf:
#             zipf.writestr(f"Rent_Data_Export_{date_str}.xlsx", excel_stream.getvalue())
#         zip_stream.seek(0)
#         zip_filename = f"Rent_Data_Archive_{date_str}.zip"
#         response = StreamingResponse(
#             iter([zip_stream.getvalue()]),
#             media_type="application/zip"
#         )
#         response.headers["Content-Disposition"] = f'attachment; filename="{zip_filename}"'
#         return response

#     elif format == "csv":
#         # Convert receipts to CSV
#         stream = io.StringIO()
#         if receipts:
#             writer = csv.DictWriter(stream, fieldnames=receipts[0].keys())
#             writer.writeheader()
#             writer.writerows(receipts)
#         else:
#             stream.write("No data found.")
        
#         csv_filename = f"receipts_export_{date_str}.csv"
#         response = StreamingResponse(
#             iter([stream.getvalue()]),
#             media_type="text/csv"
#         )
#         response.headers["Content-Disposition"] = f'attachment; filename="{csv_filename}"'
#         return response

#     raise HTTPException(status_code=400, detail="Unsupported format. Use 'xlsx', 'zip', or 'csv'.")

# def parse_excel_bytes(file_bytes, filename):
#     wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
#     if "Tenant_Profile" not in wb.sheetnames or "Rent_Receipts" not in wb.sheetnames:
#         raise ValueError(f"File '{filename}' is missing required sheets 'Tenant_Profile' and/or 'Rent_Receipts'.")

#     ws_profile = wb["Tenant_Profile"]
#     ws_receipts = wb["Rent_Receipts"]

#     p_headers = [str(cell.value).strip() if cell.value else f"Col{i}" for i, cell in enumerate(ws_profile[1])]
#     r_headers = [str(cell.value).strip() if cell.value else f"Col{i}" for i, cell in enumerate(ws_receipts[1])]

#     tenants_dict = {}
#     for row in ws_profile.iter_rows(min_row=2, values_only=True):
#         if not row[0]:
#             continue
#         row_dict = {p_headers[i]: (str(val).strip() if val is not None else "") for i, val in enumerate(row)}
#         t_id = row_dict.get("tenantId", "")
#         if t_id:
#             tenants_dict[t_id] = {"profile": row_dict, "receipts": []}

#     for row in ws_receipts.iter_rows(min_row=2, values_only=True):
#         if not row[0]:
#             continue
#         row_dict = {r_headers[i]: (str(val).strip() if val is not None else "") for i, val in enumerate(row)}
#         t_id = row_dict.get("tenantId", "")
#         if t_id in tenants_dict:
#             tenants_dict[t_id]["receipts"].append(row_dict)

#     return tenants_dict

# @router.post(Routes.ADMINAPISYNCIMPORTPREVIEW, name=Names.IMPORTPREVIEWDATA)
# async def import_preview_data(files: List[UploadFile] = File(...)):
#     preview_data = {}
#     try:
#         for file in files:
#             content = await file.read()
#             if file.filename.endswith('.zip'):
#                 with zipfile.ZipFile(io.BytesIO(content)) as z:
#                     for zip_info in z.infolist():
#                         if zip_info.filename.endswith('.xlsx'):
#                             with z.open(zip_info) as f:
#                                 preview_data[zip_info.filename] = parse_excel_bytes(f.read(), zip_info.filename)
#             elif file.filename.endswith('.xlsx'):
#                 preview_data[file.filename] = parse_excel_bytes(content, file.filename)
#             else:
#                 raise HTTPException(status_code=400, detail="Only .xlsx or .zip files are supported.")
#         return {"status": "success", "files": preview_data}
#     except HTTPException:
#         raise
#     except Exception as e:
#         raise HTTPException(status_code=400, detail=str(e))


# def _parse_excel_date(val: str) -> str:
#     """Parse Excel date strings/serials and format as 'dd MMMM yyyy'."""
#     if not val or not str(val).strip():
#         return ""
#     val_str = str(val).strip()

#     # Try ISO format: 2026-06-01 00:00:00 or 2026-06-01T00:00:00
#     for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
#         try:
#             dt = datetime.datetime.strptime(val_str, fmt)
#             return dt.strftime("%d %B %Y")  # e.g., "01 June 2026"
#         except ValueError:
#             continue

#     # Try Excel serial number
#     try:
#         serial = float(val_str)
#         # Excel epoch is 1899-12-30 (with the infamous 1900 leap year bug)
#         excel_epoch = datetime.datetime(1899, 12, 30)
#         dt = excel_epoch + datetime.timedelta(days=serial)
#         return dt.strftime("%d %B %Y")
#     except (ValueError, OverflowError):
#         pass

#     # Return original if we can't parse
#     return val_str


# def _parse_month_date(val: str) -> str:
#     """Parse Excel month value to 'Month Year' format."""
#     if not val or not str(val).strip():
#         return ""
#     val_str = str(val).strip()

#     for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
#         try:
#             dt = datetime.datetime.strptime(val_str, fmt)
#             return dt.strftime("%B %Y")  # e.g., "June 2026"
#         except ValueError:
#             continue

#     # Try Excel serial
#     try:
#         serial = float(val_str)
#         excel_epoch = datetime.datetime(1899, 12, 30)
#         dt = excel_epoch + datetime.timedelta(days=serial)
#         return dt.strftime("%B %Y")
#     except (ValueError, OverflowError):
#         pass

#     return val_str

# from typing import Dict, Literal

# VALID_TENANT_STATUSES = {"Active", "Inactive", "Archived"}

# def normalize_tenant_status(value: str | None, default: str = "Active") -> str:
#     candidate = str(value or "").strip().title()
#     return candidate if candidate in VALID_TENANT_STATUSES else default

# @router.post(Routes.ADMINAPISYNCIMPORTEXECUTE, name=Names.IMPORTEXECUTEDATA)
# async def import_execute_data(
#     background_tasks: BackgroundTasks,
#     files: List[UploadFile] = File(...),
#     selectedtargets: Optional[str] = Form(None),
#     selectedTargets: Optional[str] = Form(None),
#     targetstatuses: Optional[str] = Form(None),
# ):
#     # Accept either casing
#     targets = selectedtargets or selectedTargets or ""
#     if not targets:
#         raise HTTPException(status_code=400, detail="selectedtargets is required")
    
#     try:
#         selected_list = json.loads(targets)
#     except json.JSONDecodeError:
#         raise HTTPException(status_code=400, detail="Invalid JSON in selectedtargets")

#     if not isinstance(selected_list, list):
#         raise HTTPException(status_code=400, detail="selectedtargets must be a JSON array")

#     try:
#         status_overrides: Dict[str, str] = json.loads(targetstatuses or "{}")
#     except json.JSONDecodeError:
#         raise HTTPException(status_code=400, detail="Invalid JSON in targetstatuses.")

#     if not isinstance(status_overrides, dict):
#         raise HTTPException(status_code=400, detail="targetstatuses must be a JSON object.")

#     if not selected_list:
#         raise HTTPException(status_code=400, detail="No tenants selected for import.")

#     sys_tenants = load_tenants()
#     sys_receipts = get_all_receipts()

#     # Schedule backup BEFORE processing (non-blocking)
#     background_tasks.add_task(create_full_backup, tag="pre_import_excel")

#     # Track results for reporting
#     imported_tenants = []
#     imported_receipts = 0
#     skipped_targets = set(selected_list)  # Will remove matched ones

#     def execute_import_for_file(file_bytes, filename):
#         nonlocal imported_receipts
#         parsed_data = parse_excel_bytes(file_bytes, filename)

#         for t_id, t_data in parsed_data.items():
#             target_key = f"{filename}::{t_id}"
#             if target_key not in selected_list:
#                 continue

#             skipped_targets.discard(target_key)  # Mark as processed

#             p = t_data["profile"]
#             t_name = p.get("tenantName", "").strip()
#             if not t_name:
#                 continue

#             t = next((x for x in sys_tenants if x.name.lower() == t_name.lower()), None)
#             is_new = False
#             if not t:
#                 t = Tenant(name=t_name, phone=p.get("Phone", ""), rent=0.0, water=0.0, electricityRate=0.0)
#                 is_new = True

#             t.phone = p.get("Phone", t.phone)
#             t.email = p.get("Email", getattr(t, 'email', ''))
#             t.company = p.get("Company", getattr(t, 'company', ''))
#             t.address = p.get("Address", getattr(t, 'address', ''))
#             t.roomNumber = p.get("Room", getattr(t, 'roomNumber', ''))
#             t.meterId = p.get("meterId", getattr(t, "meterId", ""))

#             if not getattr(t, "viewToken", ""):
#                 import uuid
#                 t.viewToken = str(uuid.uuid4())

#             plain_pin = str(p.get("PIN") or "").strip()

#             pin_changed = False
#             hashed_pin = None
#             encrypted_pin = None

#             if plain_pin:
#                 validate_tenantPin(plain_pin)
#                 hashed_pin = hash_pin(plain_pin)
#                 encrypted_pin = encrypt_admin_view_pin(plain_pin)
#                 t.tenantPin = hashed_pin
#                 pin_changed = True

#             t.rent = float(p.get("Rent", t.rent) or 0.0)
#             t.water = float(p.get("Water", t.water) or 0.0)
#             t.electricityRate = float(p.get("electricityRate", t.electricityRate) or 0.0)
#             t.additionalPersonCharge = float(p.get("additionalPersonRate", getattr(t, 'additionalPersonCharge', 0.0)) or 0.0)
#             t.defaulttankWaterCharge = float(p.get("tankWater", getattr(t, 'defaulttankWaterCharge', 0.0)) or 0.0)
            
#             excel_status = normalize_tenant_status(
#                 p.get("Status"),
#                 getattr(t, "status", "Active") if not is_new else "Active",
#             )
#             requested_status = status_overrides.get(target_key)
#             t.status = normalize_tenant_status(requested_status, excel_status)

#             if is_new:
#                 tenantId = add_tenant(t)
#                 t.id = tenantId
#                 sys_tenants.append(t)

#                 if pin_changed:
#                     now = datetime.datetime.utcnow().isoformat()
#                     with get_conn() as conn:
#                         conn.execute(
#                             """
#                             INSERT INTO tenantPin_history
#                             (tenantId, pin_hash, changed_at)
#                             VALUES (?, ?, ?)
#                             """,
#                             (tenantId, hashed_pin, now)
#                         )
#                         conn.execute(
#                             """
#                             INSERT OR REPLACE INTO tenantPin_admin_store
#                             (tenantId, encrypted_pin, updated_at)
#                             VALUES (?, ?, ?)
#                             """,
#                             (tenantId, encrypted_pin, now)
#                         )
#                         conn.commit()
#             else:
#                 update_tenant(t)

#                 if pin_changed:
#                     now = datetime.datetime.utcnow().isoformat()
#                     with get_conn() as conn:
#                         conn.execute(
#                             """
#                             INSERT INTO tenantPin_history
#                             (tenantId, pin_hash, changed_at)
#                             VALUES (?, ?, ?)
#                             """,
#                             (t.id, hashed_pin, now)
#                         )
#                         conn.execute(
#                             """
#                             INSERT OR REPLACE INTO tenantPin_admin_store
#                             (tenantId, encrypted_pin, updated_at)
#                             VALUES (?, ?, ?)
#                             """,
#                             (t.id, encrypted_pin, now)
#                         )
#                         conn.commit()
#                     revoke_all_tenant_sessions(t.id)

#             imported_tenants.append({
#                 "target": target_key,
#                 "tenantId": t.id,
#                 "tenantName": t.name,
#                 "status": t.status,
#             })

#             for r in t_data["receipts"]:
#                 billNo = r.get("BillNo", "").strip()
#                 if not billNo:
#                     continue
#                 sys_r = next((x for x in sys_receipts if x.get("Bill") == billNo), None)

#                 # FIX #4, #5: Parse Excel dates to proper format
#                 raw_date = r.get("Date", "")
#                 raw_month = r.get("Month", "")

#                 data = {
#                     "Bill": billNo,
#                     "Date": _parse_excel_date(raw_date),
#                     "Month": _parse_month_date(raw_month),
#                     "Tenant": t_name,
#                     "Previous": float(r.get("Previous", 0) or 0),
#                     "Current": float(r.get("Current", 0) or 0),
#                     "Units": float(r.get("Units", 0) or 0),
#                     "Rent": float(r.get("Rent", 0) or 0),
#                     "Additional": float(r.get("Additional", 0) or 0),
#                     "Water": float(r.get("Water", 0) or 0),
#                     "tankWater": float(r.get("tankWater", 0) or 0),
#                     "Electricity": float(r.get("Electricity", 0) or 0),
#                     "MaintenanceCharge": float(r.get("Maintenance", 0) or 0),
#                     "MaintenanceDesc": r.get("MaintenanceDesc", ""),
#                     "previousArrears": float(r.get("Arrears", 0) or 0),
#                     # FIX #7: Ensure amountReceived is float
#                     "amountReceived": float(r.get("amountReceived", 0) or 0),
#                     "Total": float(r.get("Total", 0) or 0),
#                     "paymentStatus": r.get("paymentStatus", "PENDING"),
#                     "Status": r.get("receiptStatus", "ACTIVE"),
#                     "Receipt_Version": 8,
#                     "Generated_By": "Import"
#                 }
#                 if sys_r:
#                     sys_r.update(data)
#                 else:
#                     sys_receipts.append(data)
#                     imported_receipts += 1

#     # ── Process uploaded files ──
#     temp_files_to_cleanup = []

#     try:
#         for file in files:
#             content = await file.read()

#             # Collect for cleanup
#             temp_files_to_cleanup.append(file)

#             if file.filename.endswith(".zip"):
#                 with zipfile.ZipFile(io.BytesIO(content)) as z:
#                     for zip_info in z.infolist():
#                         if zip_info.filename.endswith(".xlsx"):
#                             with z.open(zip_info) as f:
#                                 execute_import_for_file(f.read(), zip_info.filename)

#             elif file.filename.endswith(".xlsx"):
#                 execute_import_for_file(content, file.filename)

#         from app.services.billing_service import save_all_receipts
#         save_all_receipts(sys_receipts)

#         # Build response message
#         msg_parts = [f"Import completed successfully."]
#         msg_parts.append(f"Tenants: {len(imported_tenants)} processed.")
#         msg_parts.append(f"Receipts: {imported_receipts} imported/updated.")

#         if skipped_targets:
#             msg_parts.append(f"Warning: {len(skipped_targets)} selected target(s) not found in files.")

#         return {
#             "status": "success",
#             "message": " ".join(msg_parts),
#             "tenants": len(imported_tenants),
#             "receipts": imported_receipts,
#             "imported_tenants": imported_tenants,
#             "unmatched_targets": list(skipped_targets) if skipped_targets else []
#         }

#     except HTTPException:
#         raise
#     except Exception as e:
#         raise HTTPException(status_code=400, detail=f"Import failed: {str(e)}")

#     finally:
#         # FIX #2: Cleanup all collected files, not just last loop variable
#         for tf in temp_files_to_cleanup:
#             try:
#                 await tf.close()
#             except Exception:
#                 pass

#             try:
#                 temp_path = getattr(tf.file, "name", None)
#                 if isinstance(temp_path, str) and os.path.isfile(temp_path):
#                     os.remove(temp_path)
#             except Exception:
#                 pass


# @router.get(Routes.ADMINAPIBILLINGARCHIVEDATA)
# async def get_archive_data():
#     tenants = load_tenants(include_archived=True)
#     archived_tenants = [tenant for tenant in tenants if tenant.status == "Archived"]
#     archived_names = {tenant.name for tenant in archived_tenants}

#     receipts = get_all_receipts(include_archived_tenants=True)
#     archived_receipts = [
#         receipt for receipt in receipts
#         if receipt.get("Status") == "ARCHIVED"
#         or receipt.get("Tenant") in archived_names
#     ]

#     return {
#         "tenants": archived_tenants,
#         "receipts": archived_receipts,
#     }

# if __name__ == "__main__":
#     sys_conf = config.get("system", {})
#     server_host = sys_conf["server"]["host"]
#     server_port = sys_conf["server"]["port"]
#     is_debug = sys_conf["server"]["debug"]

#     try:
#         s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
#         s.connect(('10.255.255.255', 1))
#         local_ip = s.getsockname()[0]
#     except Exception:
#         local_ip = '127.0.0.1'
#     finally:
#         s.close()

#     print(f"\n{'='*50}")
#     print(f" {sys_conf['app']['title']} is starting...")
#     print(f"{'='*50}")
#     print(f" [Local]:   http://127.0.0.1:{server_port}")
#     print(f" [Network]: http://{local_ip}:{server_port}")
#     print(f" [Note]:    Do NOT click the {server_host} link below")
#     print(f"{'='*50}\n")

#     uvicorn.run(
#         "app:app",
#         host=server_host,
#         port=server_port,
#         reload=is_debug,
#         proxy_headers=True,
#         forwarded_allow_ips="*",
#         access_log=True
#     )